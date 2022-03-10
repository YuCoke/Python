import requests
import json
import xlrd
from openpyxl import Workbook


# 从execle中获取查询地址列表
def user_file():

    # 输入文件路径
    file_name = input("请输入文件地址名称:")

    # 打开文件
    work = xlrd.open_workbook(file_name)
    print(work.sheet_names())
    sheet1 = work.sheets()[0]
    rows = sheet1.nrows
    print("共有", sheet1.ncols, "列")
    # 输入要使用的数据列
    f_name = sheet1.col_values(0)

    return f_name


# 根据地址名称查询经纬度
def calc_ll(name):
    total = "https://api.map.baidu.com/geocoding/v3/?address="
    body = "&output=json&ak=jX1d0dXTWgmCPQQfOyPQwQqCooM1dX7l&callback=showLocation"
    url = total + name + body
    data = {
        'address': name
    }
    address = requests.post(url, data=data).text.replace("'", '"').replace('/ ', '/')[27:-1]
    jsonValue = json.loads(address)
    if 'result' in jsonValue:
        # print(jsonValue['result']['location']['lng'])
        return [name, str(jsonValue['result']['location']['lng']), str(jsonValue['result']['location']['lat'])]

    else:
        return ''


def main():
    name_list = user_file()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '地址'
    ws['B1'] = 'lng'
    ws['C1'] = 'lat'
    num = int(0)
    i = int(0)
    for name in name_list:
        address = calc_ll(name)
        if address != '':
            num = num + 1
            # print(address)
            ws.append(address)

        else:
            i = i+1
            ws.append(name)

    wb.save('address.xlsx')
    print("成功获取数据：", num)
    print("失败的数据：", i)
    wb.close()


if __name__ == "__main__":
    main()
